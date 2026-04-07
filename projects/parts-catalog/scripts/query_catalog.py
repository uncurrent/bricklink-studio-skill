#!/usr/bin/env python3
"""
LEGO Parts Catalog — Query & Export Tool
==========================================
Run example queries against the catalog database and export results to Excel.

Usage:
    python3 query_catalog.py                    # Show example stats
    python3 query_catalog.py --export-all       # Export full catalog to Excel
    python3 query_catalog.py --search "2x4"     # Search parts by name
    python3 query_catalog.py --part 3001        # Show details for a specific part
    python3 query_catalog.py --rarest           # Show rarest part+color combos
    python3 query_catalog.py --color "Dark Red"  # Parts available in a specific color
"""

import argparse
import sqlite3
import sys
from pathlib import Path

DB_PATH = Path(__file__).parent / "lego_catalog.db"

# Optional: openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def get_db():
    if not DB_PATH.exists():
        print(f"Error: Database not found at {DB_PATH}")
        print("Run `python3 build_catalog.py` first.")
        sys.exit(1)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


# ─── Queries ──────────────────────────────────────────────────────────────────

def show_summary(conn):
    """Print overall database summary."""
    cur = conn.cursor()

    print("\n╔══════════════════════════════════════════╗")
    print("║   LEGO Parts Catalog — Summary           ║")
    print("╚══════════════════════════════════════════╝\n")

    stats = {
        "Parts": "SELECT COUNT(*) FROM parts",
        "Colors": "SELECT COUNT(*) FROM colors",
        "Part+Color combinations": "SELECT COUNT(*) FROM part_color_stats",
        "Sets": "SELECT COUNT(*) FROM sets",
        "LEGO Element IDs": "SELECT COUNT(*) FROM elements",
        "Themes": "SELECT COUNT(*) FROM themes",
    }
    for label, sql in stats.items():
        count = cur.execute(sql).fetchone()[0]
        print(f"  {label:30s} {count:>10,}")

    # Top 10 most common parts (appear in most sets)
    print("\n  ── Top 10 most common parts (by set count) ──")
    for row in cur.execute("""
        SELECT ps.part_num, p.name, ps.num_colors, ps.num_sets, ps.total_quantity
        FROM part_stats ps
        JOIN parts p ON ps.part_num = p.part_num
        ORDER BY ps.num_sets DESC
        LIMIT 10
    """):
        print(f"  {row['part_num']:12s} {row['name'][:40]:40s}  "
              f"colors={row['num_colors']:3d}  sets={row['num_sets']:5d}  qty={row['total_quantity']:8,}")

    # Top 10 rarest part+color combos (in fewest sets, min 1)
    print("\n  ── Top 10 rarest part+color combos (by set count) ──")
    for row in cur.execute("""
        SELECT pcs.part_num, p.name, c.name AS color_name, c.rgb,
               pcs.num_sets, pcs.total_quantity, pcs.first_year, pcs.last_year
        FROM part_color_stats pcs
        JOIN parts p ON pcs.part_num = p.part_num
        JOIN colors c ON pcs.color_id = c.id
        WHERE pcs.num_sets = 1
        ORDER BY pcs.total_quantity ASC
        LIMIT 10
    """):
        print(f"  {row['part_num']:12s} {row['name'][:30]:30s}  "
              f"#{row['rgb']}  {row['color_name'][:20]:20s}  "
              f"sets={row['num_sets']}  qty={row['total_quantity']}  ({row['first_year']})")

    # Color distribution
    print("\n  ── Top 15 most-used colors (by total pieces across all sets) ──")
    for row in cur.execute("""
        SELECT c.name, c.rgb, COUNT(DISTINCT pcs.part_num) AS num_parts,
               SUM(pcs.num_sets) AS total_set_appearances,
               SUM(pcs.total_quantity) AS total_qty
        FROM part_color_stats pcs
        JOIN colors c ON pcs.color_id = c.id
        GROUP BY pcs.color_id
        ORDER BY total_qty DESC
        LIMIT 15
    """):
        print(f"  #{row['rgb']}  {row['name'][:25]:25s}  "
              f"parts={row['num_parts']:6,}  total_qty={row['total_qty']:10,}")


def search_parts(conn, query):
    """Search parts by name."""
    print(f"\n  Searching for '{query}'...\n")
    cur = conn.cursor()
    rows = cur.execute("""
        SELECT p.part_num, p.name, pc.name AS category, p.part_material,
               COALESCE(ps.num_colors, 0) AS num_colors,
               COALESCE(ps.num_sets, 0) AS num_sets
        FROM parts p
        JOIN part_categories pc ON p.part_cat_id = pc.id
        LEFT JOIN part_stats ps ON p.part_num = ps.part_num
        WHERE p.name LIKE ? OR p.part_num LIKE ?
        ORDER BY COALESCE(ps.num_sets, 0) DESC
        LIMIT 50
    """, (f"%{query}%", f"%{query}%"))

    for row in rows:
        print(f"  {row['part_num']:15s} {row['name'][:50]:50s}  "
              f"[{row['category'][:20]}]  colors={row['num_colors']}  sets={row['num_sets']}")


def part_details(conn, part_num):
    """Show full details for a specific part."""
    cur = conn.cursor()

    part = cur.execute("""
        SELECT p.*, pc.name AS category_name
        FROM parts p
        JOIN part_categories pc ON p.part_cat_id = pc.id
        WHERE p.part_num = ?
    """, (part_num,)).fetchone()

    if not part:
        print(f"  Part '{part_num}' not found.")
        return

    print(f"\n  ── Part: {part['part_num']} ──")
    print(f"  Name:     {part['name']}")
    print(f"  Category: {part['category_name']}")
    print(f"  Material: {part['part_material']}")

    # Element IDs (LEGO's system)
    elements = cur.execute("""
        SELECT element_id, design_id, c.name AS color_name
        FROM elements e
        JOIN colors c ON e.color_id = c.id
        WHERE e.part_num = ?
        ORDER BY c.name
    """, (part_num,)).fetchall()

    if elements:
        print(f"\n  LEGO Element IDs ({len(elements)}):")
        for el in elements[:20]:  # show first 20
            design = f"  (Design ID: {el['design_id']})" if el['design_id'] else ""
            print(f"    {el['element_id']:12s}  {el['color_name']}{design}")
        if len(elements) > 20:
            print(f"    ... and {len(elements) - 20} more")

    # Alternate / related parts
    rels = cur.execute("""
        SELECT rel_type, child_part_num, parent_part_num
        FROM part_relationships
        WHERE child_part_num = ? OR parent_part_num = ?
    """, (part_num, part_num)).fetchall()

    if rels:
        rel_labels = {"A": "Alternate", "M": "Mold variant", "P": "Print", "B": "Sub-part", "T": "Pattern"}
        print(f"\n  Related parts ({len(rels)}):")
        for rel in rels:
            other = rel['parent_part_num'] if rel['child_part_num'] == part_num else rel['child_part_num']
            label = rel_labels.get(rel['rel_type'], rel['rel_type'])
            print(f"    {label:15s}  → {other}")

    # Available colors with frequency
    print(f"\n  Available colors:")
    colors = cur.execute("""
        SELECT c.name, c.rgb, c.is_trans, pcs.num_sets, pcs.total_quantity,
               pcs.first_year, pcs.last_year
        FROM part_color_stats pcs
        JOIN colors c ON pcs.color_id = c.id
        WHERE pcs.part_num = ?
        ORDER BY pcs.num_sets DESC
    """, (part_num,)).fetchall()

    for c in colors:
        trans = " (trans)" if c['is_trans'] else ""
        print(f"    #{c['rgb']}  {c['name'][:25]:25s}{trans:8s}  "
              f"sets={c['num_sets']:4d}  qty={c['total_quantity']:6,}  "
              f"({c['first_year']}-{c['last_year']})")


def show_color_parts(conn, color_name):
    """Show all parts available in a specific color."""
    cur = conn.cursor()

    color = cur.execute(
        "SELECT * FROM colors WHERE name LIKE ?", (f"%{color_name}%",)
    ).fetchone()

    if not color:
        print(f"  Color '{color_name}' not found. Available colors:")
        for row in cur.execute("SELECT name FROM colors ORDER BY name"):
            print(f"    {row['name']}")
        return

    print(f"\n  ── Parts in {color['name']} (#{color['rgb']}) ──\n")

    rows = cur.execute("""
        SELECT pcs.part_num, p.name, pcs.num_sets, pcs.total_quantity,
               pcs.first_year, pcs.last_year
        FROM part_color_stats pcs
        JOIN parts p ON pcs.part_num = p.part_num
        WHERE pcs.color_id = ?
        ORDER BY pcs.num_sets DESC
        LIMIT 50
    """, (color['id'],)).fetchall()

    for row in rows:
        print(f"  {row['part_num']:15s} {row['name'][:45]:45s}  "
              f"sets={row['num_sets']:4d}  qty={row['total_quantity']:6,}  "
              f"({row['first_year']}-{row['last_year']})")


def show_rarest(conn):
    """Show rarest part+color combos."""
    print("\n  ── Rarest part+color combos (appeared in only 1 set, 1 piece) ──\n")
    cur = conn.cursor()
    for row in cur.execute("""
        SELECT pcs.part_num, p.name, c.name AS color_name, c.rgb,
               pcs.num_sets, pcs.total_quantity, pcs.first_year,
               s.name AS set_name, s.set_num
        FROM part_color_stats pcs
        JOIN parts p ON pcs.part_num = p.part_num
        JOIN colors c ON pcs.color_id = c.id
        -- find the one set it appeared in
        JOIN inventory_parts ip ON ip.part_num = pcs.part_num AND ip.color_id = pcs.color_id
        JOIN inventories inv ON ip.inventory_id = inv.id
        JOIN sets s ON inv.set_num = s.set_num
        WHERE pcs.num_sets = 1 AND pcs.total_quantity = 1
        ORDER BY pcs.first_year DESC
        LIMIT 30
    """):
        print(f"  {row['part_num']:12s}  {row['name'][:30]:30s}  "
              f"#{row['rgb']} {row['color_name'][:18]:18s}  "
              f"← {row['set_num']} {row['set_name'][:30]}  ({row['first_year']})")


# ─── Excel Export ─────────────────────────────────────────────────────────────

def export_to_excel(conn):
    """Export full catalog to Excel with formatting."""
    if not HAS_OPENPYXL:
        print("  Error: openpyxl not installed. Run: pip3 install openpyxl")
        sys.exit(1)

    export_path = Path(__file__).parent / "lego_catalog_export.xlsx"
    print(f"\n  Exporting to Excel: {export_path}")

    wb = openpyxl.Workbook()
    cur = conn.cursor()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")

    def make_sheet(ws, title, sql, col_widths=None):
        ws.title = title
        rows = cur.execute(sql).fetchall()
        if not rows:
            return

        # Headers
        keys = rows[0].keys()
        for col, key in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(keys, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[key])

        # Column widths
        if col_widths:
            for col, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        print(f"    {title}: {len(rows):,} rows")

    # Sheet 1: Parts with stats
    make_sheet(wb.active, "Parts Catalog", """
        SELECT p.part_num AS "Part #",
               p.name AS "Name",
               pc.name AS "Category",
               p.part_material AS "Material",
               COALESCE(ps.num_colors, 0) AS "Colors Available",
               COALESCE(ps.num_sets, 0) AS "Sets Count",
               COALESCE(ps.total_quantity, 0) AS "Total Quantity",
               COALESCE(ps.first_year, '') AS "First Year",
               COALESCE(ps.last_year, '') AS "Last Year"
        FROM parts p
        JOIN part_categories pc ON p.part_cat_id = pc.id
        LEFT JOIN part_stats ps ON p.part_num = ps.part_num
        ORDER BY COALESCE(ps.num_sets, 0) DESC
    """, [15, 45, 20, 12, 15, 12, 14, 10, 10])

    # Sheet 2: Part+Color combos with frequency
    ws2 = wb.create_sheet()
    make_sheet(ws2, "Part × Color", """
        SELECT pcs.part_num AS "Part #",
               p.name AS "Part Name",
               c.name AS "Color",
               '#' || c.rgb AS "RGB",
               CASE WHEN c.is_trans THEN 'Yes' ELSE 'No' END AS "Transparent",
               pcs.num_sets AS "Sets Count",
               pcs.total_quantity AS "Total Qty",
               pcs.first_year AS "First Year",
               pcs.last_year AS "Last Year"
        FROM part_color_stats pcs
        JOIN parts p ON pcs.part_num = p.part_num
        JOIN colors c ON pcs.color_id = c.id
        ORDER BY pcs.num_sets DESC
        LIMIT 500000
    """, [15, 40, 20, 10, 12, 12, 12, 10, 10])

    # Sheet 3: Colors overview
    ws3 = wb.create_sheet()
    make_sheet(ws3, "Colors", """
        SELECT c.id AS "Color ID",
               c.name AS "Color Name",
               '#' || c.rgb AS "RGB",
               CASE WHEN c.is_trans THEN 'Yes' ELSE 'No' END AS "Transparent",
               COUNT(DISTINCT pcs.part_num) AS "Parts Available",
               SUM(pcs.total_quantity) AS "Total Pieces in Sets"
        FROM colors c
        LEFT JOIN part_color_stats pcs ON c.id = pcs.color_id
        GROUP BY c.id
        ORDER BY SUM(pcs.total_quantity) DESC
    """, [10, 25, 10, 12, 15, 20])

    # Sheet 4: Element ID mapping (BrickLink ↔ LEGO)
    ws4 = wb.create_sheet()
    make_sheet(ws4, "Element ID Map", """
        SELECT e.element_id AS "LEGO Element ID",
               e.part_num AS "Part # (BrickLink/Rebrickable)",
               e.design_id AS "LEGO Design ID",
               p.name AS "Part Name",
               c.name AS "Color"
        FROM elements e
        JOIN parts p ON e.part_num = p.part_num
        JOIN colors c ON e.color_id = c.id
        ORDER BY e.part_num, c.name
    """, [15, 30, 15, 40, 20])

    wb.save(str(export_path))
    print(f"\n  ✓ Exported to {export_path}")


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Query the LEGO Parts Catalog")
    parser.add_argument("--search", help="Search parts by name or number")
    parser.add_argument("--part", help="Show details for a specific part number")
    parser.add_argument("--color", help="Show parts available in a specific color")
    parser.add_argument("--rarest", action="store_true", help="Show rarest part+color combos")
    parser.add_argument("--export-all", action="store_true", help="Export full catalog to Excel")
    args = parser.parse_args()

    conn = get_db()

    if args.search:
        search_parts(conn, args.search)
    elif args.part:
        part_details(conn, args.part)
    elif args.color:
        show_color_parts(conn, args.color)
    elif args.rarest:
        show_rarest(conn)
    elif args.export_all:
        export_to_excel(conn)
    else:
        show_summary(conn)

    conn.close()
